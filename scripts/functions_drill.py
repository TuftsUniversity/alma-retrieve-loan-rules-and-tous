from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementClickInterceptedException, StaleElementReferenceException, TimeoutException

import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

def login(driver, username, password):
    """Login to Alma"""
    username_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'username')))
    username_field.send_keys(username)

    password_field = driver.find_element(By.ID, 'password')
    password_field.send_keys(password)

    password_field.submit()
    return driver

def safe_find_element(driver, by, value, retries=3):
    """Find element with retries, and refresh the page if retries fail."""
    for attempt in range(retries):
        try:
            return WebDriverWait(driver, 25).until(EC.visibility_of_element_located((by, value)))
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None
def safe_find_element_text(driver, by, value, retries=3):
    """Find element with retries for StaleElementReferenceException"""

    for attempt in range(retries):
        try:
            element =  WebDriverWait(driver, 20).until(EC.visibility_of_element_located((by, value)))
            return element.text
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
def click_element_with_retry(driver, by, value, retries=3, wait_time=10):
    """Click an element with retry to handle stale elements dynamically."""
    for attempt in range(retries):
        try:
            # Wait for element to be present
            element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((by, value)))

            # Wait until element is clickable
            element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((by, value)))
             # Click the element
            element.click()
            return
        except ElementClickInterceptedException:
            print(f"Attempt {attempt + 1} of {retries}: Click intercepted by overlay, handling...")

            # Check if an overlay is blocking and remove it
            try:
                overlay = driver.find_element(By.CLASS_NAME, "mask")  # Adjust class name if needed
                driver.execute_script("arguments[0].remove();", overlay)  # Remove the overlay
                print("Overlay detected and removed.")
            except:
                print("No overlay found.")

            time.sleep(2)  # Give time for changes before retrying
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
                print(f"Failed to locate element: {value} after {retries} retries.")

        except:

            return None  # Return None instead of throwing an exception
def get_table_html_with_retry(driver, by, value, table, retries=3):



    # --- Extract Loan Policy Data ---
        # loan_tab = safe_find_element(driver, By.ID, "A_NAV_LINK_touTypeloan_span")


        
        

        # # Extract policy table
        # loan_policy_table_html = get_table_html_with_retry(
        #     driver, By.ID, "TABLE_DATA_policiesList"
        # )

        # loan_policy_df = pd.read_html(loan_policy_table_html)[0]
    """Retries getting table HTML to handle stale element issues."""
    for attempt in range(retries):
        # try:
        if table == "loan":
            #print("got into table == loan")
            click_element_with_retry(driver, By.ID, "A_NAV_LINK_touTypeloan_span")
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: "Is Loanable" in d.find_element(by, value).get_attribute("outerHTML")
                )
                #("Request tab content loaded.")
            except TimeoutException:
                print("⚠️ Timeout waiting for 'Is Requestable' content in request tab")
                raise
            #fulfillment_unit_name = safe_find_element_text(driver, By.XPATH, "//div[contains(@class, 'row ') and .//span[contains(text(), 'Fulfillment Unit Name')]]//a")

            table_element = get_html(driver, by, value) # WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, value)))
                # If it's a string of HTML, parse it
            
            policy_df = get_dataframe(table_element)

            policy_df = policy_df[['Policy Type', 'Policy Description']]
            policy_df = policy_df.set_index('Policy Type').T
            policy_series = policy_df.squeeze(axis=0)

            policy_dict = policy_series.to_dict()

            #print("Parsed DataFrame:")
            #print(policy_df)

            # Check for required columns
            # if "Is Loanable" not in policy_df["Policy Type"].values:
            #     raise Exception
        
            #else:
            tou_name = safe_find_element_text(
            driver,
            By.XPATH,
            "//div[contains(@class, 'row ') and .//span[contains(text(), 'Terms Of Use Name')]]//a",
            )

            fulfillment_rule_name = safe_find_element_text(
            driver,
            By.XPATH,
            "//div[contains(@class, 'row ') and .//span[contains(text(), 'Fulfillment Unit Rule')]]//a",
            )

            fulfillment_unit_name = safe_find_element_text(driver, By.XPATH, "//div[contains(@class, 'row ') and .//span[contains(text(), 'Fulfillment Unit Name')]]//a")

            return [fulfillment_rule_name, tou_name, policy_dict, fulfillment_unit_name]
    
        elif table == "request":
            #print("got into table == request")
            click_element_with_retry(driver, By.ID, "A_NAV_LINK_touTyperequest_span")
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: "Is Requestable" in d.find_element(by, value).get_attribute("outerHTML")
                )
                #print("Request tab content loaded.")
            except TimeoutException:
                print("⚠️ Timeout waiting for 'Is Requestable' content in request tab")
                raise

            table_element = get_html(driver, by, value) # WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, value)))
            # If it's a string of HTML, parse it
            # if isinstance(table_element, str):
            #     try:
            #         dfs = pd.read_html(table_element)
            #         if dfs:
            #             policy_df = dfs[0]
            #         else:
            #             raise ValueError("No tables found in HTML.")
            #     except Exception as e:
            #         raise RuntimeError(f"Failed to parse HTML table: {e}")
            # elif isinstance(table_element, pd.DataFrame):
            #     policy_df = table_element
            # else:
            #     raise TypeError(f"Unexpected input type for table_element: {type(table_element)}")

            # policy_df = policy_df.fillna("")

            #print("Parsed DataFrame:")
            

            policy_df = get_dataframe(table_element)

            policy_df = policy_df[['Policy Type', 'Policy Description']]
            policy_df = policy_df.set_index('Policy Type').T
            policy_series = policy_df.squeeze(axis=0)

            policy_dict = policy_series.to_dict()


            # Check for required columns
            # if "Is Requestable" not in policy_df["Policy Type"].values:
            #     raise Exception
            #else:
            tou_name = safe_find_element_text(
            driver,
            By.XPATH,
            "//div[contains(@class, 'row ') and .//span[contains(text(), 'Terms Of Use Name')]]//a",
            )

            fulfillment_rule_name = safe_find_element_text(
            driver,
            By.XPATH,
            "//div[contains(@class, 'row ') and .//span[contains(text(), 'Fulfillment Unit Rule')]]//a",
            )

            fulfillment_unit_name = safe_find_element_text(driver, By.XPATH, "//div[contains(@class, 'row ') and .//span[contains(text(), 'Fulfillment Unit Name')]]//a")

            return [fulfillment_rule_name, tou_name, policy_dict, fulfillment_unit_name]
    

        else:
            print("got into Exception.    did not find loan nav link")
            raise Exception
        
            
            
        # except StaleElementReferenceException:
        #     print(f"Attempt {attempt + 1} of {retries}: Table element stale, retrying...")
        #     time.sleep(2)

        # except Exception:
        #     print("Wrong table type provided by interface")
        #     time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception

def get_dataframe(table_element):
    if isinstance(table_element, str):
        try:
            dfs = pd.read_html(table_element)
            if dfs:
                policy_df = dfs[0]
            else:
                raise ValueError("No tables found in HTML.")
        except Exception as e:
            raise RuntimeError(f"Failed to parse HTML table: {e}")
    elif isinstance(table_element, pd.DataFrame):
        policy_df = table_element
    else:
        raise TypeError(f"Unexpected input type for table_element: {type(table_element)}")

    policy_df = policy_df.fillna("")

    return policy_df
def get_html(driver, by, value, retries=3):
    """Retries getting table HTML to handle stale element issues."""
    for attempt in range(retries):
        try:
            table_element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((by, value)))
            return table_element.get_attribute('outerHTML')
        except StaleElementReferenceException:
            print(f"Attempt {attempt + 1} of {retries}: Table element stale, retrying...")
            time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
def send_keys_with_retry(driver, by, value, text, retries=3, wait_time=10):
    """Send keys to an element with retry to handle stale element issues."""
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((by, value))
            )
            element.clear()
            element.send_keys(text)
            return
        except:
            print(f"Attempt {attempt + 1} of {retries}: Stale element reference, retrying...")
            time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None  # Return None instead of throwing an exception
import pandas as pd
from openpyxl import load_workbook

def append_to_excel(file_path, buffer):
    """Append buffer data to Excel file in batches or force flush when needed"""

    buffer_df = pd.DataFrame(buffer)
    #print(file_path)
    # Ensure existing data is loaded if file exists
    if os.path.exists(file_path):
        if os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                buffer_df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    else:
        buffer_df.to_excel(file_path, index=False)

    print(f"🔹 Wrote {len(buffer_df)} records to {file_path} (Total size: {os.path.getsize(file_path) / 1024:.2f} KB)")

    buffer.clear()  # Clear buffer after writing

import os
import pandas as pd
from openpyxl import load_workbook

def write_buffer_to_excel(buffer, thread_id, output_dir):
    if not buffer:
        return

    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, f"output_thread_{thread_id}.xlsx")
    df = pd.DataFrame(buffer)

    try:
        if os.path.exists(file_path):
            # Load workbook to determine start row
            book = load_workbook(file_path)
            sheet = book.active
            start_row = sheet.max_row

            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, header=False, startrow=start_row)
        else:
            df.to_excel(file_path, index=False)

        print(f"✅ Thread-{thread_id} wrote {len(df)} rows to {file_path}")
        buffer.clear()
    except Exception as e:
        print(f"❌ Error in write_buffer_to_excel for Thread-{thread_id}: {e}")


def merge_excel_files(num_threads, output_dir):
    all_data = []
    for i in range(num_threads):
        file_path = os.path.join(output_dir, f"output_thread_{i}.xlsx")
        if os.path.exists(file_path):
            print(f"✅ Including: {file_path}")
            df = pd.read_excel(file_path, engine="openpyxl")
            all_data.append(df)
        else:
            print(f"⚠️ File not found: {file_path}")

    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)
        print(f"📁 Merged {len(all_data)} files. Total rows: {len(final_df)}")
        return final_df
    else:
        print("❌ No files to merge.")
        return pd.DataFrame()

def retrieve_current_row_index(output_dir):
    if not os.path.isdir(output_dir) or not os.listdir(output_dir):
        return 0, 0

    output_files = [f for f in os.listdir(output_dir) if f.startswith("output_thread_") and f.endswith(".xlsx")]
    num_existing_threads = len(output_files)

    combined_df = merge_excel_files(num_existing_threads, output_dir)
    if not combined_df.empty:
        return len(combined_df), num_existing_threads
    else:
        return 0, num_existing_threads
def highlight_unique_values(file_path, output_path):
    # Load the spreadsheet into a pandas DataFrame
    df = pd.read_excel(file_path, engine='openpyxl')

    # Ensure column G exists
    if len(df.columns) < 7:  # Column G is the 8th column (0-indexed)
        raise ValueError("Column G does not exist in the spreadsheet.")

    df = df.sort_values(by=["Location", "Fulfillment Rule (Loan)"])
    # Get unique values in column G
    unique_values = df.iloc[:, 6].dropna().unique()  # Column H (0-indexed)
    color_map = {}

    # Generate unique colors for each unique value in Column H
    for i, value in enumerate(unique_values):
        # Generate color codes in ARGB format (8-character hex string)
        red = (100 + (i * 50) % 256) % 256
        green = (150 + (i * 30) % 256) % 256
        blue = (200 + (i * 70) % 256) % 256
        color_map[value] = f"FF{red:02X}{green:02X}{blue:02X}"

    # Load workbook and active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Iterate through column H and apply fill
    for row in range(2, sheet.max_row + 1):  # Skip header (row 1)
        cell = sheet[f'G{row}']
        value = cell.value
        if value in color_map:
            fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
            cell.fill = fill

    # Save the updated workbook

    workbook.save(file_path)
    print(f"File saves: {file_path}")