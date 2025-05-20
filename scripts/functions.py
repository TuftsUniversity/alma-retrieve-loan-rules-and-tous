#!/usr/bin/env python3

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec


import pandas as pd
import numpy as np


from selenium.webdriver.support.ui import Select

import time
import re
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    StaleElementReferenceException,
    TimeoutException,
)


def safe_find_element(driver, by, value, retries=3):
    """Find element with retries, and refresh the page if retries fail."""
    for attempt in range(retries):
        try:
            return WebDriverWait(driver, 25).until(
                EC.visibility_of_element_located((by, value))
            )
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None


def safe_find_element_text(driver, by, value, retries=3):
    """Find element with retries for StaleElementReferenceException"""
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((by, value))
            )
            return element.text
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()  # Refresh the page on final attempt
                time.sleep(5)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None


def click_element_with_retry(driver, by, value, retries=3, wait_time=10):
    """Click an element with retry to handle stale elements dynamically."""
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((by, value))
            )
            element = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((by, value))
            )
            element.click()
            return
        except ElementClickInterceptedException:
            print(
                f"Attempt {attempt + 1} of {retries}: Click intercepted by overlay, handling..."
            )
            try:
                overlay = driver.find_element(By.CLASS_NAME, "mask")
                driver.execute_script("arguments[0].remove();", overlay)
                print("Overlay detected and removed.")
            except:
                print("No overlay found.")
            time.sleep(2)
        except StaleElementReferenceException:
            print(f"Retry {attempt + 1} of {retries}: Element stale, retrying...")
            time.sleep(2)
            if attempt == retries - 1:
                print("Too many stale element errors. Refreshing page...")
                driver.refresh()
                time.sleep(5)
                print(f"Failed to locate element: {value} after {retries} retries.")
        except:
            return None


def send_keys_with_retry(driver, by, value, text, retries=3, wait_time=10):
    """Send keys to an element with retry to handle stale element issues."""
    for attempt in range(retries):
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((by, value))
            )
            element.clear()
            element.send_keys(text)
            return
        except:
            print(
                f"Attempt {attempt + 1} of {retries}: Stale element reference, retrying..."
            )
            time.sleep(2)
    print(f"Failed to locate element: {value} after {retries} retries.")
    return None


def login(driver, username, password):
    element = driver.find_element(By.ID, "username")

    time.sleep(2)
    element.send_keys(username)

    element = driver.find_element(By.ID, "password")

    element.send_keys(password)

    element.send_keys(Keys.RETURN)

    return element


def navigate_to_fulfillment_units(driver, url):
    time.sleep(2)

    driver.get(
        url
        + "/ng/page;u=%2Fful%2Faction%2FpageAction.do%3FxmlFileName%3DfulfillmentUnits.fulfillment_units_list.xml&almaConfiguration%3Dtrue&pageViewMode%3DEdit&operation%3DLOAD&pageBean.orgUnitCode%3D3851&pageBean.currentUrl%3DxmlFileName%253DfulfillmentUnits.fulfillment_units_list.xml%2526almaConfiguration%253Dtrue%2526pageViewMode%253DEdit%2526operation%253DLOAD%2526pageBean.orgUnitCode%253D3851%2526resetPaginationContext%253Dtrue%2526showBackButton%253Dfalse&pageBean.navigationBackUrl%3D..%252Faction%252Fhome.do&resetPaginationContext%3Dtrue&showBackButton%3Dfalse&pageBean.ngBack%3Dtrue;ng=true"
    )

    time.sleep(6)

    html = driver.page_source
    fulfillment_unit_df = pd.read_html(html)[0]

    return fulfillment_unit_df


def get_fulfillment_unit_count(driver):
    time.sleep(3)
    count = len(driver.find_elements(By.XPATH, "//table/tbody/tr"))
    return count


def navigate_to_rules_tab_get_lists(driver, full_unit_number):
    click_element_with_retry(driver, By.XPATH, f'//*[contains(@id, "input_fulfillmentUnits_{full_unit_number}")]')
    click_element_with_retry(driver, By.XPATH, f'//*[contains(@id, "fulfillmentUnits_{full_unit_number}_c.ui.table.btn.edit")]/a')
    click_element_with_retry(driver, By.XPATH, "//*[@id='fulfillmentunit_editfulfillmentUnitLocations_span']/a")

    content = safe_find_element(driver, By.ID, "fulfillmentUnitLocationsList").get_attribute("outerHTML")
    locations_df = pd.read_html(content)[0]
    print("initial locations")
    with pd.option_context('display.max_columns', None):
        print(locations_df)
    # if len(locations_df) >= 20:
    #     click_element_with_retry(driver, By.ID, "navigaionSizeBarSize2")
    #     content = safe_find_element(driver, By.XPATH, '//*[contains(@id, "fulfillmentUnitLocationsList")]').get_attribute("outerHTML")
    #     locations_df = pd.read_html(content)[0]
    #     print("got into locations greater than 20")
    #     print("greater than 20 locations list")
    #     with pd.option_context('display.max_columns', None):
    #         print(locations_df)

    locations_list = locations_df["sorted ascending"].tolist()

    click_element_with_retry(driver, By.XPATH, '//*[@id="fulfillmentunit_editfulfillmentUnitRules_span"]/a')
    time.sleep(2)
    rules_df = pd.read_html(driver.page_source)[0]

    # print(rules_df)
    print("\n\n\n")
    # print(locations_df)

    return [rules_df, locations_list]


###################################
####    This gets the value of the
####    enabled slider switch to
####    see if the rule is active
def get_enabled_value(driver, y):
    element = safe_find_element(driver, By.XPATH, f"//*[contains(@id, 'ROW_{y}_COL_activeImage')]/div")
    aria_label = element.get_attribute("aria-label")

    if re.search(r"\sactive", aria_label):
        return "Active"
    elif re.search(r"inactive", aria_label):
        return "Inactive"
    return "Unknown"

def navigate_to_loan_rule(driver, y):
    button_xpath = f'//*[@id="input_rules_{y}"]'
    click_element_with_retry(driver, By.XPATH, button_xpath)

    edit_xpath = f'//*[@id="ROW_ACTION_rules_{y}_c.ui.table.btn.edit"]/a'
    click_element_with_retry(driver, By.XPATH, edit_xpath)

###################################
####    This function retrieves the parameter
####    table from the loan rule and compresses the
####    potentially multi-row table into a single, key-value
####    type list and joins that into a string to enter into
####    the Parameters field of the row assoicated with this loan rule
def get_parameter_list(driver):
    time.sleep(1)
    try:
        parameter_df = pd.read_html(driver.page_source)[0]
        parameter_list = [
            f"{row['Name']} {row['Operator']} {row['Value']}"
            for _, row in parameter_df.iterrows()
        ]
        return "; ".join(parameter_list)
    except:
        return ""

def navigate_to_tou(driver):
    click_element_with_retry(driver, By.XPATH, '//*[@id="uiconfiguration_rule_detailsview_tou"]')

###################################
####    This ingests the TOU table,
####    sets the Policy Name as the dataframe index
####    and the "type" as value, and squeezes this datafram into
####    a series, that can be appeneded as additional columns in the
####    dataframe row for the associated loan rule
def get_tou_as_series(driver):
    time.sleep(1)
    try:
        tou_df = pd.read_html(driver.page_source)[0][["Policy Name", "Policy Type"]]
        return tou_df.set_index("Policy Type").squeeze()
    except:
        return pd.Series()
    

def write_buffer_to_excel(buffer, thread_id, output_dir):
    # if not buffer:
    #     return

    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, f"output_thread_{thread_id}.xlsx")
    df = pd.DataFrame(buffer)

    try:
        if os.path.exists(file_path):
            # Load workbook to determine start row
            book = load_workbook(file_path)
            sheet = book.active
            start_row = sheet.max_row

            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, header=False, startrow=start_row)
        else:
            df.to_excel(file_path, index=False)

        print(f"✅ Thread-{thread_id} wrote {len(df)} rows to {file_path}")
        # buffer.clear()
    except Exception as e:
        print(f"❌ Error in write_buffer_to_excel for Thread-{thread_id}: {e}")


import os
import glob
import pandas as pd

def merge_excel_files(output_dir, output_file):
    """Merge all Excel files in the output directory into a single file."""
    all_data = []

    # Get all .xlsx files in the directory
    file_paths = glob.glob(os.path.join(output_dir, "*.xlsx"))

    if not file_paths:
        print("❌ No Excel files found to merge.")
        return

    for file_path in file_paths:
        try:
            print(f"✅ Including: {file_path}")
            df = pd.read_excel(file_path, engine="openpyxl")
            all_data.append(df)
        except Exception as e:
            print(f"⚠️ Failed to read {file_path}: {e}")

    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)

        if os.path.exists(output_file):
            os.remove(output_file)

        final_df.to_excel(output_file, index=False)
        print(f"📁 Merged {len(file_paths)} files. Final output: {output_file} ({os.path.getsize(output_file)/1024:.2f} KB)")
    else:
        print("❌ No valid data found to merge.")