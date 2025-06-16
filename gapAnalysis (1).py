#!/usr/bin/env python3
import os
import pandas as pd
import re
import time
import sys
import json
sys.path.append(os.path.relpath('config/'))
import secrets_local
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
# Add the full path to "scripts" based on current script location
current_dir = os.path.dirname(os.path.abspath(__file__))
scripts_dir = os.path.join(current_dir, "scripts")
if scripts_dir not in sys.path:
    sys.path.append(scripts_dir)
from functions import *


############################################################################
############################################################################
####
####    Title:  gapAnalysis.py
####    Author: Henry Steele, Senion Systems Librarian, Library Technology Services, Tufts University
####    Purpose
####        ingest three files from bulk loan rule testing, loan rule report export, and mapping of old item
####        policies to new item policies, and apply a column in the two former reports that contain the new item policy
####        field that will be written to the old item policy
####        
####        Then analyze the bulk loan rule testing file to see if applying this new item policy, which usually
####        aims to simplify the organization of loan rules, will have any unexecpted consequences
####    Input:
####        Ingest three files via file picker
####         - Bulk_Checkout_Request_Results - Formatted.xlsx from the bulk loan rule tester
####         - loan rule report from getLoanRules.py
####         - mapping of old item policies to new item policies
####    Method:
####       - after ingesting 3 files, sort bulk loan rule tester exporter formatted by (in order)
####          - location
####          - new item policy
####          - user group
####       - then group these by those fields, iteratively for each group
####       - the old item policies in this group will likely differ within each of these, but what this script
####         seeks to identify is within these groups, are there more than one loan rule and TOU/request rule and TOU
####         because this would be unexpected.  It may not be unwanted, but at least should be identified

FORMATTED_OUTPUT_FILE = 'Potential Issues with Item Policy Application to Loan - Gap Analysis.xlsx'
OUTPUT_DIR = 'Output'
#inputFilenameBulkTestingFormatted = askopenfilename(title="Select Excel file Bulk_Checkout_Request_Results - Formatted.xlsx")
#inputFilenameAnalytics = askopenfilename(title="Select Excel file containing loan rules containing extant Item Policy/Location combinations")
#inputFilenameMapping = askopenfilename(title="Select Excel file containing mapping of old item policy to new item policy")
inputFilenameBulkTestingFormatted = "Bulk_Checkout_Request_Results - Formatted.xlsx"
inputFilenameMapping = "input/Item Policy Change/approved/Item_Policy_Loan_Mapping.xlsx"
wb = load_workbook(inputFilenameBulkTestingFormatted, data_only=True)
ws = wb['Sheet1']


visible_rows = []
for row in ws.iter_rows():  # Removed values_only=True
    if not ws.row_dimensions[row[0].row].hidden:
        visible_rows.append([cell.value for cell in row])

header = [cell for cell in visible_rows[0]]
data_rows = [[cell for cell in row] for row in visible_rows[1:]]
df_tester = pd.DataFrame(data_rows, columns=header)

df_tester = pd.DataFrame(visible_rows[1:], columns=visible_rows[0])

#df_analytics = pd.read_excel(inputFilenameAnalytics, engine="openpyxl", dtype='str')
df_mapping = pd.read_excel(inputFilenameMapping, engine="openpyxl", dtype='str')

df_tester = df_tester.sort_values(by=[])

''' 
make a new column in both sheets.  the easiest is going to be formatted.  loan rule report will rely on item policy being in separate column

the new column will apply the new item policy where old item policy matches item policy in other sheet
'''


#df_mapping = df_mapping['Current Item Policy','New Item Policy']

mapping_dict = {}

for index, row in df_mapping.iterrows():
    mapping_dict[row['Library Name'] + "-" + row['Current Item Policy']] = row['New item policy/Loan Length']

prefixes = ("Ginn", "HHSL", "Hirsh", "Music", "SMFA", "Vet")
#df_tester['Location'] = df_tester['Location'].apply(
    #lambda x: x if any(x.strip().startswith(p) for p in prefixes) else "Tisch " + x
#)

# print(json.dumps(mapping_dict))

df_tester['Library'] = df_tester['Location'].apply(lambda x: x.replace("Reserves", "Library"))
df_tester['Library'] = df_tester['Library'].apply(lambda x: re.sub(r"^([\S]+)\s?(Library)?.+$", r"\1", x))
df_tester['Library Name-Item Policy'] = df_tester['Library'] + "-" + df_tester['Item Policy']


#df_analytics['Library'] = df_analytics['Location Name'].apply(lambda x: x.replace("Reserves", "Library"))
#df_analytics['Library'] = df_analytics['Library'].apply(lambda x: re.sub(r"^([\S]+)(Library)?.+$", r"\1", x))
#df_analytics['Library Name-Item Policy'] = df_analytics['Library'] + "-" + df_analytics['Item Policy']

missing_keys_list = []

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', 10)


pd.set_option('display.width', None)        # Prevents line wrapping
pd.set_option('display.max_colwidth', None) # Ensures full content in each cell is shown


# print("df_tester from Bulk Loan Rules Testing")

# print(df_tester)
# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', 10)


# pd.set_option('display.width', None)        # Prevents line wrapping
# pd.set_option('display.max_colwidth', None) # Ensures full content in each cell is shown

# print("df analytics")
# print(df_analytics)

def missing_keys(x):
    try:
        return mapping_dict[x]
    except KeyError:
        missing_keys_list.append(x)
        return None  # or some default value like 'UNKNOWN'

df_tester['New Item Policy'] = df_tester['Library Name-Item Policy'].apply(missing_keys)




#df_tester = df_tester.drop('Library Name-Item Policy', axis=1)
                
                           
non_extant_library_item_policy_combinations = []
unique_library_list = df_tester['Library'].unique()

unique_item_policy_list = df_mapping['New item policy/Loan Length'].unique()

unique_user_group_list = df_tester['User Group'].unique()
check_df = pd.DataFrame(columns=df_tester.columns)

for library in unique_library_list:
    for item_policy in unique_item_policy_list:
        
        #if library + "-" + item_policy not in df_tester['Library Name-Item Policy'].values or library + "-" + item_policy not in df_analytics['Library Name-Item Policy'].values:
            #print("Non-extant library/item policy combination: " + library + "-" + item_policy)
            non_extant_library_item_policy_combinations.append(library + "-" + item_policy)
            #continue
        
        
    for user_group in unique_user_group_list:
    
        df_tester_group = df_tester.copy()
        df_tester_group = df_tester[(df_tester['Library'] == library) & (df_tester['New Item Policy'] == item_policy) & (df_tester['User Group'] == user_group)]
            
            # df_tester_group = df_tester_group.reset_index()
            # for location in df_tester_group['Location'].unique():
            # # pd.options.display.max_rows = None
            #     tester_location_subgroup_df = df_tester_group.copy()

        tester_location_subgroup_df = tester_location_subgroup_df[tester_location_subgroup_df['Location'] == location]
            # pd.options.display.max_columns = None
            # print(df_tester_group)
        if len(tester_location_subgroup_df) > 1 and len(tester_location_subgroup_df['Fulfillment Rule (Loan)'].unique()) > 1:  
                            
            check_df = pd.concat([check_df, tester_location_subgroup_df])


check_df.to_excel('Potential Issues with Item Policy Application to Loan - Gap Analysis.xlsx', index=False)
 
highlight_unique_values(FORMATTED_OUTPUT_FILE, OUTPUT_DIR)
print("All threads complete.")
# Filter out keys that are also in non_extant_library_item_policy_combinations
filtered_missing_keys = [key for key in missing_keys_list if key not in non_extant_library_item_policy_combinations]

# Prepare the string for writing
missing_keys_file_string = "\n".join(filtered_missing_keys)

# Write to file
with open("Missing Library-Item Policy Mappings.txt", "w+") as missing_keys_file:
    missing_keys_file.write(missing_keys_file_string)

# Prepare and write the non-extant combinations
non_extant_mappings_skip_string = "\n".join(non_extant_library_item_policy_combinations)
with open("Non-Extant Mappings String.  Ignore.txt", "w+") as non_extant_mappings_file:
    non_extant_mappings_file.write(non_extant_mappings_skip_string)

missing_keys_file.close()

